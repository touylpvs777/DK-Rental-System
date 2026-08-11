"""
End-to-end validation script — DK CRM
Tests Customer, Lead, Activity Log, and Reports workflows against the live API.
"""
import csv as csvmod
import io
import sys
import time
from datetime import date

import httpx
import openpyxl

BASE = "http://localhost:8000/api/v1"
TS   = str(int(time.time()))

GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
RESET  = "\033[0m"
BOLD   = "\033[1m"

passed, failed, warned = [], [], []

def ok(msg):
    passed.append(msg)
    print(f"  {GREEN}PASS{RESET}  {msg}")

def fail(msg, detail=""):
    failed.append(msg)
    suffix = f"\n         detail: {detail}" if detail else ""
    print(f"  {RED}FAIL{RESET}  {msg}{suffix}")

def warn(msg):
    warned.append(msg)
    print(f"  {YELLOW}WARN{RESET}  {msg}")

def section(title):
    print(f"\n{BOLD}{CYAN}{'='*65}{RESET}")
    print(f"{BOLD}{CYAN}  {title}{RESET}")
    print(f"{BOLD}{CYAN}{'='*65}{RESET}")

client = httpx.Client(timeout=15)

def get(path, **kw):   return client.get(f"{BASE}{path}", **kw)
def post(path, **kw):  return client.post(f"{BASE}{path}", **kw)
def put(path, **kw):   return client.put(f"{BASE}{path}", **kw)
def patch(path, **kw): return client.patch(f"{BASE}{path}", **kw)
def delete(path, **kw):return client.delete(f"{BASE}{path}", **kw)

# ─────────────────────────────────────────────────────────────────────────────
section("0 · Authentication")
# ─────────────────────────────────────────────────────────────────────────────
r = client.post(f"{BASE}/auth/login", json={"username": "admin", "password": "Admin@123"})
if r.status_code == 200:
    token = r.json()["access_token"]
    client.headers["Authorization"] = f"Bearer {token}"
    ok("POST /auth/login → 200, access_token returned")
else:
    fail(f"Login → {r.status_code}", r.text)
    sys.exit(1)

r = get("/users/me")
if r.status_code == 200:
    me = r.json()
    ok(f"GET /users/me → 200  (user={me['username']}, superuser={me['is_superuser']})")
else:
    fail(f"GET /users/me → {r.status_code}", r.text)

# ─────────────────────────────────────────────────────────────────────────────
section("1 · Customer Workflow")
# ─────────────────────────────────────────────────────────────────────────────
c_email = f"e2e_{TS}@test.com"

# CREATE
r = post("/customers/", json={
    "first_name": "E2E", "last_name": "Customer",
    "email": c_email, "phone": "+1 555 000 1234",
    "company": "E2E Corp", "status": "prospect",
    "notes": "Created by automated E2E test",
})
if r.status_code == 201:
    cust = r.json(); cid = cust["id"]
    ok(f"POST /customers/ → 201  id={cid}")
    ok("  notes field returned") if cust.get("notes") == "Created by automated E2E test" \
        else fail("  notes field missing/wrong", str(cust.get("notes")))
    ok("  status=prospect") if cust.get("status") == "prospect" \
        else fail("  status wrong", cust.get("status"))
else:
    fail(f"POST /customers/ → {r.status_code}", r.text); cid = None

# READ back
if cid:
    r = get(f"/customers/{cid}")
    ok(f"GET /customers/{cid} → 200") if r.status_code == 200 \
        else fail(f"GET /customers/{cid} → {r.status_code}", r.text)

# LIST — customer appears
r = get("/customers/")
if r.status_code == 200:
    items = r.json()
    ok(f"GET /customers/ → 200  ({len(items)} total)")
    ok("  new customer in list") if cid and any(c["id"] == cid for c in items) \
        else fail("  new customer not in list")
else:
    fail(f"GET /customers/ → {r.status_code}", r.text)

# EDIT — update name, status, notes  (customer uses PATCH)
if cid:
    r = patch(f"/customers/{cid}", json={
        "first_name": "E2E-Edited", "last_name": "Customer",
        "email": c_email, "status": "active", "notes": "Updated by E2E",
    })
    if r.status_code == 200:
        u = r.json()
        ok(f"PATCH /customers/{cid} → 200")
        ok("  first_name updated") if u.get("first_name") == "E2E-Edited" \
            else fail("  first_name not updated", u.get("first_name"))
        ok("  status changed to active") if u.get("status") == "active" \
            else fail("  status not changed", u.get("status"))
        ok("  notes updated") if u.get("notes") == "Updated by E2E" \
            else fail("  notes not updated", u.get("notes"))
    else:
        fail(f"PATCH /customers/{cid} → {r.status_code}", r.text)

# CHURNED status (new value added in BUG-02 fix)
if cid:
    r = patch(f"/customers/{cid}", json={"status": "churned"})
    ok("  status=churned accepted and stored") \
        if r.status_code == 200 and r.json().get("status") == "churned" \
        else fail("  status=churned rejected", f"{r.status_code} {r.text[:100]}")

# DUPLICATE email → 409
r = post("/customers/", json={"first_name": "Dup", "last_name": "Test", "email": c_email})
ok("  duplicate email → 409 Conflict") if r.status_code == 409 \
    else fail(f"  duplicate email → {r.status_code} (expected 409)", r.text[:100])

# BLANK name → 422
r = post("/customers/", json={"first_name": "", "last_name": "Test"})
ok("  blank first_name → 422") if r.status_code == 422 \
    else fail(f"  blank first_name → {r.status_code} (expected 422)", r.text[:100])

# DELETE
if cid:
    r = delete(f"/customers/{cid}")
    if r.status_code == 204:
        ok(f"DELETE /customers/{cid} → 204")
        r2 = get(f"/customers/{cid}")
        ok("  deleted customer → 404") if r2.status_code == 404 \
            else fail(f"  deleted customer returns {r2.status_code}")
    else:
        fail(f"DELETE /customers/{cid} → {r.status_code}", r.text)

# ─────────────────────────────────────────────────────────────────────────────
section("2 · Lead Workflow")
# ─────────────────────────────────────────────────────────────────────────────

# Create a customer to link leads to
r = post("/customers/", json={
    "first_name": "Lead", "last_name": "Owner",
    "email": f"lo_{TS}@test.com", "status": "active",
})
lo_cid = r.json()["id"] if r.status_code == 201 else None
if lo_cid:
    ok(f"Created lead-owner customer id={lo_cid}")
else:
    warn("Could not create customer for leads")

# CREATE lead
r = post("/leads/", json={
    "title": f"E2E Lead {TS}", "description": "Automated test lead",
    "value": 50000.0, "status": "new",
    "customer_id": lo_cid, "source": "website",
})
if r.status_code == 201:
    lead = r.json(); lid = lead["id"]
    ok(f"POST /leads/ → 201  id={lid}")
    ok("  value=50000.0") if lead.get("value") == 50000.0 \
        else fail("  value wrong", str(lead.get("value")))
    ok("  status=new") if lead.get("status") == "new" \
        else fail("  status wrong", lead.get("status"))
else:
    fail(f"POST /leads/ → {r.status_code}", r.text); lid = None

# BLANK title → 422
r = post("/leads/", json={"title": ""})
ok("  blank title → 422") if r.status_code == 422 \
    else fail(f"  blank title → {r.status_code} (expected 422)", r.text[:100])

# NEGATIVE value → 422
r = post("/leads/", json={"title": "Bad", "value": -100.0})
ok("  negative value → 422") if r.status_code == 422 \
    else fail(f"  negative value → {r.status_code} (expected 422)", r.text[:100])

# READ
if lid:
    r = get(f"/leads/{lid}")
    if r.status_code == 200:
        detail = r.json()
        ok(f"GET /leads/{lid} → 200")
        ok("  customer nested in detail") \
            if detail.get("customer") and detail["customer"]["id"] == lo_cid \
            else fail("  customer relationship missing", str(detail.get("customer")))
    else:
        fail(f"GET /leads/{lid} → {r.status_code}", r.text)

# EDIT
if lid:
    r = put(f"/leads/{lid}", json={"title": f"E2E Lead {TS} EDITED", "value": 75000.0})
    if r.status_code == 200:
        ok(f"PUT /leads/{lid} → 200")
        ok("  value updated to 75000") if r.json().get("value") == 75000.0 \
            else fail("  value not updated", str(r.json().get("value")))
    else:
        fail(f"PUT /leads/{lid} → {r.status_code}", r.text)

# STATUS TRANSITIONS (full happy path) — use PUT /leads/{id} with {"status": ...}
transitions = [
    ("new",       "contacted"),
    ("contacted", "qualified"),
    ("qualified", "proposal"),
    ("proposal",  "won"),
]
if lid:
    for frm, to in transitions:
        r = put(f"/leads/{lid}", json={"status": to})
        ok(f"  {frm}→{to} accepted") \
            if r.status_code == 200 and r.json().get("status") == to \
            else fail(f"  {frm}→{to} failed", f"{r.status_code} {r.text[:80]}")

# INVALID TRANSITION (won → new should be rejected via PUT)
if lid:
    r = put(f"/leads/{lid}", json={"status": "new"})
    ok("  invalid won→new → 422 rejected") if r.status_code == 422 \
        else fail(f"  invalid transition won→new not rejected → {r.status_code}", r.text[:100])

# ADD NOTE
note_id = None
if lid:
    r = post(f"/leads/{lid}/notes", json={"content": "E2E test note"})
    if r.status_code == 201:
        note_id = r.json()["id"]
        ok(f"POST /leads/{lid}/notes → 201  note_id={note_id}")
    else:
        fail(f"POST /leads/{lid}/notes → {r.status_code}", r.text)

# BLANK note → 422
if lid:
    r = post(f"/leads/{lid}/notes", json={"content": ""})
    ok("  blank note content → 422") if r.status_code == 422 \
        else fail(f"  blank note → {r.status_code} (expected 422)", r.text[:100])

# DELETE NOTE
if lid and note_id:
    r = delete(f"/leads/{lid}/notes/{note_id}")
    ok(f"DELETE note {note_id} → 204") if r.status_code == 204 \
        else fail(f"DELETE note → {r.status_code}", r.text)

# DELETE LEAD
if lid:
    r = delete(f"/leads/{lid}")
    if r.status_code == 204:
        ok(f"DELETE /leads/{lid} → 204")
        r2 = get(f"/leads/{lid}")
        ok("  deleted lead → 404") if r2.status_code == 404 \
            else fail(f"  deleted lead returns {r2.status_code}")
    else:
        fail(f"DELETE /leads/{lid} → {r.status_code}", r.text)

if lo_cid:
    delete(f"/customers/{lo_cid}")

# ─────────────────────────────────────────────────────────────────────────────
section("3 · Activity Log")
# ─────────────────────────────────────────────────────────────────────────────

# Generate known activity: create → update → delete a customer
r = post("/customers/", json={
    "first_name": "ActivityTest", "last_name": "User",
    "email": f"activity_{TS}@test.com", "status": "prospect",
})
act_cid = r.json()["id"] if r.status_code == 201 else None
if act_cid:
    put(f"/customers/{act_cid}", json={"status": "active"})
    delete(f"/customers/{act_cid}")

time.sleep(0.4)   # let writes settle

r = get("/activity/", params={"limit": 500})
if r.status_code == 200:
    logs = r.json()
    ok(f"GET /activity/ → 200  ({len(logs)} total records)")
    actions = [e["action"] for e in logs]

    for expected in ["user_login", "customer_created", "customer_updated",
                     "customer_status_changed", "customer_deleted"]:
        ok(f"  action '{expected}' logged") if expected in actions \
            else fail(f"  action '{expected}' NOT in log")
else:
    fail(f"GET /activity/ → {r.status_code}", r.text[:100])
    logs = []

# Date filter — today (BUG-03 fix check)
today = date.today().isoformat()
r = get("/activity/", params={"from_date": today, "to_date": today, "limit": 500})
if r.status_code == 200:
    day_logs = r.json()
    ok(f"  Date filter today → 200, {len(day_logs)} records")
    # Records are stored as UTC naive datetimes; derive the actual UTC date
    # from the newest record so the filter matches regardless of local timezone
    if len(day_logs) == 0 and logs:
        utc_date = logs[0]["created_at"][:10]   # "YYYY-MM-DD" portion
        r_utc = get("/activity/", params={"from_date": utc_date, "to_date": utc_date, "limit": 500})
        if r_utc.status_code == 200 and len(r_utc.json()) > 0:
            ok(f"  BUG-03 date filter works (UTC date {utc_date} → {len(r_utc.json())} records; local date differs from UTC)")
        else:
            fail(f"  BUG-03: date filter returned 0 records even for UTC date {utc_date}")
    elif len(day_logs) > 0:
        ok("  BUG-03 date filter returns non-empty results")
else:
    fail(f"  Date filter → {r.status_code}", r.text[:100])

# /activity/me
r = get("/activity/me")
ok(f"GET /activity/me → 200  ({len(r.json())} records)") if r.status_code == 200 \
    else fail(f"GET /activity/me → {r.status_code}", r.text[:80])

# Entity history for a customer
clog = next((e for e in logs if e.get("action") == "customer_created" and e.get("entity_id")), None)
if clog:
    eid = clog["entity_id"]
    r = get(f"/activity/entity/customer/{eid}")
    ok(f"  Entity history /activity/entity/customer/{eid} → 200  ({len(r.json())} records)") \
        if r.status_code == 200 else fail(f"  Entity history → {r.status_code}", r.text[:80])

# /activity/user/{id} (superuser endpoint)
r = get(f"/activity/user/{me['id']}", params={"limit": 50})
ok(f"  /activity/user/{me['id']} → 200  ({len(r.json())} records)") if r.status_code == 200 \
    else fail(f"  /activity/user → {r.status_code}", r.text[:80])

# ─────────────────────────────────────────────────────────────────────────────
section("4 · Reports")
# ─────────────────────────────────────────────────────────────────────────────

# Seed: 2 customers, 1 won lead, 1 lost lead
def seed():
    rc1 = post("/customers/", json={"first_name": "RptA", "last_name": "C",
                                    "email": f"rpta_{TS}@test.com", "status": "active"})
    rc2 = post("/customers/", json={"first_name": "RptB", "last_name": "C",
                                    "email": f"rptb_{TS}@test.com", "status": "prospect"})
    c1 = rc1.json()["id"] if rc1.status_code == 201 else None
    c2 = rc2.json()["id"] if rc2.status_code == 201 else None

    rl1 = post("/leads/", json={"title": f"Won {TS}", "value": 12000.0,
                                 "status": "new", "customer_id": c1})
    rl2 = post("/leads/", json={"title": f"Lost {TS}", "value": 8000.0,
                                 "status": "new", "customer_id": c2})
    l1 = rl1.json()["id"] if rl1.status_code == 201 else None
    l2 = rl2.json()["id"] if rl2.status_code == 201 else None

    if l1:
        for s in ["contacted", "qualified", "proposal", "won"]:
            put(f"/leads/{l1}", json={"status": s})
    if l2:
        for s in ["contacted", "lost"]:
            put(f"/leads/{l2}", json={"status": s})
    return c1, c2, l1, l2

rc1, rc2, rl1, rl2 = seed()

# ── 4a Customer CSV
r = get("/reports/customers", params={"format": "csv"})
if r.status_code == 200:
    ok("GET /reports/customers?format=csv → 200")
    ct = r.headers.get("content-type", "")
    ok("  Content-Type: text/csv") if "text/csv" in ct \
        else fail("  wrong Content-Type", ct)
    raw = r.content.decode("utf-8-sig")
    rows = list(csvmod.DictReader(io.StringIO(raw)))
    ok(f"  {len(rows)} data rows")
    if rows:
        missing = [c for c in ["id","first_name","last_name","email","status"]
                   if c not in rows[0]]
        ok("  All expected columns present") if not missing \
            else fail("  Missing columns", str(missing))
    # Verify our seeded customers appear
    emails_in_csv = {row.get("email","") for row in rows}
    ok("  Seeded customer rpta appears in CSV") \
        if f"rpta_{TS}@test.com" in emails_in_csv \
        else fail("  Seeded customer rpta not in CSV")
else:
    fail(f"GET /reports/customers?format=csv → {r.status_code}", r.text[:120])

# ── 4b Customer Excel
r = get("/reports/customers", params={"format": "excel"})
if r.status_code == 200:
    ok("GET /reports/customers?format=excel → 200")
    ok("  Content-Type: xlsx") if "spreadsheetml" in r.headers.get("content-type","") \
        else fail("  wrong Content-Type", r.headers.get("content-type"))
    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        xrows = list(ws.iter_rows(values_only=True))
        ok(f"  Excel parses OK, {len(xrows)} rows (incl. header)")
        header_cells = [str(c).lower() for c in xrows[0] if c]
        ok("  first_name in Excel header") \
            if any("first_name" in h for h in header_cells) \
            else fail("  first_name missing in Excel header", str(xrows[0]))
        # Check freeze panes set
        ok("  freeze_panes set on A2") if str(ws.freeze_panes) == "A2" \
            else warn(f"  freeze_panes = {ws.freeze_panes} (expected A2)")
    except Exception as ex:
        fail("  Excel parse error", str(ex))
else:
    fail(f"GET /reports/customers?format=excel → {r.status_code}", r.text[:120])

# ── 4c Lead CSV
r = get("/reports/leads", params={"format": "csv"})
if r.status_code == 200:
    ok("GET /reports/leads?format=csv → 200")
    raw = r.content.decode("utf-8-sig")
    lrows = list(csvmod.DictReader(io.StringIO(raw)))
    ok(f"  {len(lrows)} lead rows in CSV")
    titles = {row.get("title","") for row in lrows}
    ok("  Seeded Won lead appears in CSV") \
        if f"Won {TS}" in titles \
        else fail(f"  Seeded 'Won {TS}' not in lead CSV", str(list(titles)[:5]))
    ok("  Seeded Lost lead appears in CSV") \
        if f"Lost {TS}" in titles \
        else fail(f"  Seeded 'Lost {TS}' not in lead CSV")
else:
    fail(f"GET /reports/leads?format=csv → {r.status_code}", r.text[:120])

# ── 4d Sales CSV (BUG-01 fix verification)
r = get("/reports/sales", params={"format": "csv"})
if r.status_code == 200:
    ok("GET /reports/sales?format=csv → 200  (BUG-01 confirmed fixed)")
    raw = r.content.decode("utf-8-sig")
    rows = list(csvmod.DictReader(io.StringIO(raw)))
    ok(f"  {len(rows)} rows in Sales CSV (detail + footer)")
    text = raw.lower()
    ok("  Summary section present") if "summary" in text \
        else warn("  No summary section in Sales CSV")
    ok("  'won' label present") if "won" in text \
        else warn("  'won' not in Sales CSV")
    ok("  'lost' label present") if "lost" in text \
        else warn("  'lost' not in Sales CSV")

    # Verify our 12000 won lead appears in the detail section
    value_rows = [row for row in rows if row.get("value","").replace(".","").lstrip("-").isdigit()
                  or (row.get("value") and row["value"].replace(".","").replace("-","").isdigit())]
    try:
        values = [float(row["value"]) for row in rows
                  if row.get("value","").strip() not in ("", "──")]
        ok(f"  Parseable value cells: {len(values)}  sum={sum(values):,.2f}")
        if 12000.0 in values:
            ok("  Won lead value 12000 appears in sales detail")
        else:
            warn(f"  12000 not found in values list: {values[:10]}")
        if 8000.0 in values:
            ok("  Lost lead value 8000 appears in sales detail")
        else:
            warn(f"  8000 not found in values: {values[:10]}")
    except Exception as ex:
        warn(f"  Could not parse sales values: {ex}")
else:
    fail(f"GET /reports/sales?format=csv → {r.status_code}", r.text[:200])

# ── 4e Sales Excel
r = get("/reports/sales", params={"format": "excel"})
if r.status_code == 200:
    ok("GET /reports/sales?format=excel → 200")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        xrows = list(ws.iter_rows(values_only=True))
        ok(f"  Sales Excel parses OK, {len(xrows)} rows")
        all_vals = [str(c) for row in xrows for c in row if c is not None]
        ok("  Summary rows in Sales Excel") \
            if any("SUMMARY" in v or "Won Revenue" in v for v in all_vals) \
            else warn("  Summary rows not found in Sales Excel")
    except Exception as ex:
        fail("  Sales Excel parse error", str(ex))
else:
    fail(f"GET /reports/sales?format=excel → {r.status_code}", r.text[:120])

# ── 4f Lead Excel
r = get("/reports/leads", params={"format": "excel"})
if r.status_code == 200:
    ok("GET /reports/leads?format=excel → 200")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        xrows = list(ws.iter_rows(values_only=True))
        ok(f"  Lead Excel parses OK, {len(xrows)} rows")
    except Exception as ex:
        fail("  Lead Excel parse error", str(ex))
else:
    fail(f"GET /reports/leads?format=excel → {r.status_code}", r.text[:120])

# ── 4g Dashboard metrics accuracy  (route is /dashboard/summary)
# Schema fields: total_customers, active_customers, total_leads, won_leads,
#                conversion_rate, win_rate, lost_rate, leads_by_source, etc.
r = get("/dashboard/summary")
if r.status_code == 200:
    dash = r.json()
    ok("GET /dashboard/summary → 200")
    for key in ["total_customers","total_leads","won_leads","conversion_rate","win_rate"]:
        ok(f"  {key} = {dash.get(key)}") if key in dash \
            else fail(f"  dashboard.{key} missing", str(list(dash.keys())))
    # Sanity: won_leads should be >= 1 since we seeded a won lead
    wl = dash.get("won_leads", 0)
    ok(f"  won_leads={wl} >= 1 (seeded won lead counted)") if wl >= 1 \
        else warn(f"  won_leads={wl} — seeded won lead may not be counted yet")
else:
    fail(f"GET /dashboard/summary → {r.status_code}", r.text[:120])

# ── 4h Report date filter
r = get("/reports/customers", params={"format": "csv", "from_date": today, "to_date": today})
ok(f"  Report date filter today → 200") if r.status_code == 200 \
    else fail(f"  Report date filter → {r.status_code}", r.text[:100])

# ─────────────────────────────────────────────────────────────────────────────
section("5 · Auth — Token Revocation (BUG-05)")
# ─────────────────────────────────────────────────────────────────────────────
temp = httpx.Client(timeout=10)
r = temp.post(f"{BASE}/auth/login", json={"username": "admin", "password": "Admin@123"})
if r.status_code == 200:
    temp_tok = r.json()["access_token"]
    temp.headers["Authorization"] = f"Bearer {temp_tok}"
    ok("Second login → 200")

    r = temp.get(f"{BASE}/users/me")
    ok("Token valid before logout → 200") if r.status_code == 200 \
        else fail(f"Token invalid before logout → {r.status_code}", r.text[:80])

    r = temp.post(f"{BASE}/auth/logout")
    ok("POST /auth/logout → 204") if r.status_code == 204 \
        else fail(f"POST /auth/logout → {r.status_code}", r.text[:120])

    time.sleep(0.1)
    r = temp.get(f"{BASE}/users/me")
    ok("Revoked token → 401 (revocation works)") if r.status_code == 401 \
        else fail(f"Revoked token accepted → {r.status_code} (expected 401)", r.text[:120])

    # Also verify logout is logged
    time.sleep(0.2)
    r2 = get("/activity/", params={"limit": 100})
    if r2.status_code == 200:
        acts = [e["action"] for e in r2.json()]
        ok("  user_logout action logged") if "user_logout" in acts \
            else fail("  user_logout not found in activity log")
else:
    fail(f"Second login → {r.status_code}", r.text[:80])
temp.close()

# ─────────────────────────────────────────────────────────────────────────────
section("6 · Edge Cases")
# ─────────────────────────────────────────────────────────────────────────────
r = get("/customers/999999")
ok("GET non-existent customer → 404") if r.status_code == 404 \
    else fail(f"GET /customers/999999 → {r.status_code}")

r = get("/leads/999999")
ok("GET non-existent lead → 404") if r.status_code == 404 \
    else fail(f"GET /leads/999999 → {r.status_code}")

r = post("/customers/", json={"first_name": "Bad", "last_name": "Email", "email": "notanemail"})
ok("Invalid email → 422") if r.status_code == 422 \
    else fail(f"Invalid email → {r.status_code}", r.text[:80])

r = post("/leads/", json={"title": "BadVal", "value": -500.0})
ok("Negative lead value → 422") if r.status_code == 422 \
    else fail(f"Negative lead value → {r.status_code}", r.text[:80])

# DELETE on already-deleted resource → 404
r = post("/customers/", json={"first_name": "Temp", "last_name": "Del",
                               "email": f"tempdel_{TS}@test.com"})
if r.status_code == 201:
    tid = r.json()["id"]
    delete(f"/customers/{tid}")
    r2 = delete(f"/customers/{tid}")
    ok("Double-delete → 404") if r2.status_code == 404 \
        else warn(f"Double-delete → {r2.status_code} (expected 404)")

# ─────────────────────────────────────────────────────────────────────────────
section("Cleanup")
# ─────────────────────────────────────────────────────────────────────────────
for res_type, ids in [("leads", [rl1, rl2]), ("customers", [rc1, rc2])]:
    for rid in ids:
        if rid:
            dr = delete(f"/{res_type}/{rid}")
            print(f"  cleaned /{res_type}/{rid}: {dr.status_code}")

client.close()

# ─────────────────────────────────────────────────────────────────────────────
section("FINAL RESULTS")
# ─────────────────────────────────────────────────────────────────────────────
total = len(passed) + len(failed) + len(warned)
score_pct = round(100 * len(passed) / total) if total else 0

print(f"\n  {GREEN}Passed  : {len(passed)}{RESET}")
print(f"  {YELLOW}Warnings: {len(warned)}{RESET}")
print(f"  {RED}Failed  : {len(failed)}{RESET}")
print(f"  Total   : {total}")
print(f"\n  Check score: {score_pct}% ({len(passed)}/{total})\n")

if failed:
    print(f"{RED}FAILURES:{RESET}")
    for f_ in failed:
        print(f"  {RED}✗{RESET}  {f_}")

if warned:
    print(f"\n{YELLOW}WARNINGS:{RESET}")
    for w in warned:
        print(f"  {YELLOW}⚠{RESET}  {w}")

sys.exit(1 if failed else 0)

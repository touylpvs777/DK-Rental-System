"""
Export helpers: convert list[dict] → StreamingResponse (CSV or Excel).
No database access — pure I/O utilities.
"""
import csv
import io
from datetime import datetime

from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def _fmt(value: object) -> object:
    """Stringify enums and datetimes so writers never see raw Python objects."""
    if hasattr(value, "value"):          # str enum
        return value.value
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


# ── CSV ───────────────────────────────────────────────────────────────────────

def csv_response(data: list[dict], filename: str) -> StreamingResponse:
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(
            output, fieldnames=list(data[0].keys()), lineterminator="\n",
            extrasaction="ignore",
        )
        writer.writeheader()
        for row in data:
            writer.writerow({k: _fmt(v) for k, v in row.items()})

    # UTF-8 BOM so Excel opens the file without a wizard
    content = ("﻿" + output.getvalue()).encode("utf-8")
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Excel ─────────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="2563EB")   # blue
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def excel_response(
    data: list[dict],
    filename: str,
    sheet_name: str = "Report",
) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if data:
        headers = list(data[0].keys())

        # Header row
        ws.append(headers)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        # Data rows
        for row in data:
            ws.append([_fmt(v) for v in row.values()])

        # Auto-fit column widths (capped at 50)
        for col_idx, header in enumerate(headers, start=1):
            col_values = [str(_fmt(row[header]) or "") for row in data]
            max_len = max(len(header), max((len(v) for v in col_values), default=0))
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"   # keep header visible on scroll

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

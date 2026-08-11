"""
Excel export/import for Purchase Orders (openpyxl).

Export produces a fully unlocked .xlsx laid out like a standard printed PO
(company header, vendor block, ship-to, line-item table, totals, and an
Issued/Reviewed/Approved-By signature row) so a user can open it, edit
quantities/prices/descriptions freely, and re-upload it.

Import re-reads that same fixed layout back — this is a round-trip format
for our own export, not a general-purpose "any spreadsheet" importer like
`InventoryImportService` (which does header-alias matching for arbitrary
CSV/Excel uploads). The PO is identified by its `po_number` cell; the sheet's
item rows fully replace the PO's existing items (an edited sheet describes
the new complete item set, not a diff).
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.purchase_order import PurchaseOrder
from app.schemas.po_excel import POExcelImportRowError, POExcelImportResult

logger = logging.getLogger(__name__)

_THIN = Side(style="thin", color="999999")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL_FONT = Font(bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=16)
_LABEL_FONT = Font(bold=True, size=9, color="666666")

# Fixed layout rows/columns so import can read the same file back deterministically.
_ITEM_HEADER_ROW = 12
_ITEM_START_ROW = 13
_ITEM_COLUMNS = ("item_code", "description", "unit", "quantity_ordered", "unit_cost", "line_total")
_ITEM_HEADER_LABELS = ("Part No", "Description", "U/M", "Quantity", "Unit Price", "Total")
_MAX_ITEM_ROWS = 500  # sane upper bound so a corrupt/huge sheet can't hang the import


def _fmt_money(n: float) -> float:
    return round(n, 2)


class POExcelService:
    """Stateless — pure (bytes in/out) formatting and parsing logic, no DB access."""

    # ── Export ───────────────────────────────────────────────────────────────

    def export(self, po: PurchaseOrder) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Order"
        # No `ws.protection.sheet = True` anywhere in this module — the sheet
        # is unprotected by default, i.e. every cell stays freely editable.

        for col, width in zip("ABCDEF", (16, 40, 10, 12, 14, 14)):
            ws.column_dimensions[col].width = width

        ws.merge_cells("A1:C1")
        ws["A1"] = "DK Service"
        ws["A1"].font = _TITLE_FONT
        ws.merge_cells("D1:F1")
        ws["D1"] = "PURCHASE ORDER"
        ws["D1"].font = _TITLE_FONT
        ws["D1"].alignment = Alignment(horizontal="right")

        ws["A2"] = po.vendor_address or ""
        ws["D2"] = "Date"
        ws["E2"] = po.order_date.isoformat() if po.order_date else ""
        ws["A3"] = po.vendor_contact or ""
        ws["D3"] = "PO Number"
        ws["E3"] = po.po_number
        for cell in ("D2", "D3"):
            ws[cell].font = _LABEL_FONT

        ws["A5"] = "VENDOR"
        ws["A5"].font = _LABEL_FONT
        ws["A6"] = po.vendor

        ws["A8"] = "SHIP TO"
        ws["A8"].font = _LABEL_FONT
        ws["A9"] = (po.warehouse.name if po.warehouse else "") or ""
        ws["A10"] = (po.warehouse.address if po.warehouse and po.warehouse.address else "") or ""

        for i, label in enumerate(_ITEM_HEADER_LABELS):
            cell = ws.cell(row=_ITEM_HEADER_ROW, column=i + 1, value=label)
            cell.font = _HEADER_FILL_FONT
            cell.border = _BOX
            cell.alignment = Alignment(horizontal="center")

        row = _ITEM_START_ROW
        for item in po.items:
            values = (
                item.item_code or "", item.description or "", item.unit or "",
                item.quantity_ordered, item.unit_cost, item.line_total,
            )
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = _BOX
                if col_idx >= 4:
                    cell.number_format = "#,##0.00"
            row += 1
        last_item_row = max(row - 1, _ITEM_START_ROW)

        totals_row = last_item_row + 2
        for label, value in (
            ("Subtotal", po.subtotal),
            (f"Tax ({po.tax_rate:g}%)", po.tax_amount),
            ("Grand Total", po.total_amount),
        ):
            ws.cell(row=totals_row, column=5, value=label).font = _LABEL_FONT
            c = ws.cell(row=totals_row, column=6, value=_fmt_money(value))
            c.number_format = "#,##0.00"
            totals_row += 1

        sig_row = totals_row + 2
        for col_idx, role_label in enumerate(("Issued By", "Reviewed By", "Approved By")):
            letter = get_column_letter(col_idx * 2 + 1)
            ws[f"{letter}{sig_row}"] = "_" * 20
            ws[f"{letter}{sig_row + 1}"] = role_label
            ws[f"{letter}{sig_row + 1}"].font = _LABEL_FONT

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Import ───────────────────────────────────────────────────────────────

    def parse(self, content: bytes) -> tuple[str, list[dict], list[POExcelImportRowError]]:
        """Returns (po_number, item_rows, row_errors). Raises HTTPException for
        file-level failures (unreadable/wrong layout); per-row problems are
        collected into `row_errors` instead of aborting the whole import."""
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws: Worksheet = wb.active
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot read Excel file: {exc}")

        po_number = ws["E3"].value
        if not po_number or not str(po_number).strip():
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail="Cell E3 (PO Number) is empty — this file isn't a PO export, or the layout was altered.",
            )
        po_number = str(po_number).strip()

        header = [ws.cell(row=_ITEM_HEADER_ROW, column=i + 1).value for i in range(len(_ITEM_HEADER_LABELS))]
        expected = list(_ITEM_HEADER_LABELS)
        if [str(h).strip() if h else "" for h in header] != expected:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=f"Item table header row ({_ITEM_HEADER_ROW}) doesn't match the expected columns {expected} "
                       "— rows/columns must not be inserted or deleted above the item table.",
            )

        rows: list[dict] = []
        errors: list[POExcelImportRowError] = []
        blank_streak = 0
        for offset in range(_MAX_ITEM_ROWS):
            row_number = _ITEM_START_ROW + offset
            values = [ws.cell(row=row_number, column=i + 1).value for i in range(len(_ITEM_COLUMNS))]
            if all(v is None or str(v).strip() == "" for v in values):
                blank_streak += 1
                if blank_streak >= 3:  # a few consecutive blanks past the last row = end of table
                    break
                continue
            blank_streak = 0

            item_code, description, unit, qty_raw, cost_raw, _total_raw = values
            try:
                if not (description and str(description).strip()):
                    raise ValueError("Description is required")
                quantity = float(qty_raw) if qty_raw not in (None, "") else 0.0
                unit_cost = float(cost_raw) if cost_raw not in (None, "") else 0.0
                if quantity <= 0:
                    raise ValueError(f"Quantity must be greater than 0 (got {qty_raw!r})")
                if unit_cost < 0:
                    raise ValueError(f"Unit price cannot be negative (got {cost_raw!r})")
            except (TypeError, ValueError) as exc:
                errors.append(POExcelImportRowError(row_number=row_number, error_message=str(exc)))
                continue

            rows.append({
                "item_code": (str(item_code).strip() if item_code else None),
                "description": str(description).strip(),
                "unit": (str(unit).strip() if unit else None),
                "quantity_ordered": quantity,
                "unit_cost": unit_cost,
                "line_total": _fmt_money(quantity * unit_cost),
            })

        return po_number, rows, errors

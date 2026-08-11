"""
Excel Import Service for the Product Catalog.

Two-phase flow:
  1. preview()  — parse + validate, store results in ImportJob.preview_data (no product writes)
  2. execute()  — replay preview_data rows, upsert products, update job status

Sheet detection uses keyword matching against Lao/English sheet names.
"""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any

from fastapi import HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.import_job import ImportError, ImportJob, ImportStatus
from app.repositories.import_repository import ImportRepository
from app.repositories.product_repository import ProductRepository
from app.schemas.import_job import (
    ImportExecuteResponse,
    ImportPreviewResponse,
    PreviewError,
    PreviewProduct,
    PreviewSpec,
    SheetPreview,
)
from app.services.brand_service import BrandService
from app.services.category_service import CategoryService
from app.services.notification_service import ADMIN_ROLE, NotificationService, resolve_actor_name
from app.services.product_service import ProductService
from app.schemas.product import ProductCreate, SpecCreate

logger = logging.getLogger(__name__)

# ── Sheet handler mapping ─────────────────────────────────────────────────────
# Keys are substrings (lowercased) found in sheet names; values are handler keys.
_SHEET_MAP: list[tuple[str, str]] = [
    ("jungheinrich",      "jungheinrich_forklift"),
    ("mitsubishi",        "mitsubishi_forklift"),
    # Lao keywords
    ("ອຸປະກອນ",           "accessories"),
    ("ສາງ",               "warehouse"),
    ("ດູດຝຸ່ນ",            "vacuum"),
    ("ເຊັດຄັດ",            "floor_cleaner"),
    ("ບໍລິການ",            "brands"),
    # English fallbacks
    ("vacuum",            "vacuum"),
    ("scrubber",          "floor_cleaner"),
    ("sweeper",           "floor_cleaner"),
    ("racking",           "warehouse"),
    ("warehouse",         "warehouse"),
    ("accessories",       "accessories"),
    ("spare",             "accessories"),
    ("service",           "brands"),
]

_IGNORED_SHEETS = {"ຂໍ້ມູນບໍລິສັດ", "ໂຄງສ້າງການຈັດຕັ້ງ", "ລູກຄ້າ"}


# ── Domain-specific category paths ───────────────────────────────────────────
_CATEGORY_PATHS: dict[str, tuple[str, str, str | None]] = {
    "jungheinrich_forklift": ("Material Handling Equipment", "Forklifts", "Electric Counterbalance"),
    "mitsubishi_diesel":     ("Material Handling Equipment", "Forklifts", "Diesel Counterbalance"),
    "mitsubishi_lpg":        ("Material Handling Equipment", "Forklifts", "LPG Counterbalance"),
    "mitsubishi_electric3":  ("Material Handling Equipment", "Forklifts", "Electric 3-Wheel"),
    "mitsubishi_electric4":  ("Material Handling Equipment", "Forklifts", "Electric 4-Wheel"),
    "mitsubishi_other":      ("Material Handling Equipment", "Forklifts", "Other Forklifts"),
    "accessories":           ("Material Handling Equipment", "Attachments & Accessories", None),
    "warehouse":             ("Warehouse Solutions", "Warehouse Equipment", None),
    "racking":               ("Warehouse Solutions", "Racking & Shelving", None),
    "vacuum_commercial":     ("Cleaning Equipment", "Vacuum Cleaners", "Commercial Vacuum"),
    "vacuum_industrial":     ("Cleaning Equipment", "Vacuum Cleaners", "Industrial Vacuum"),
    "scrubber_dryer":        ("Cleaning Equipment", "Floor Cleaning", "Scrubber Dryers"),
    "sweeper":               ("Cleaning Equipment", "Floor Cleaning", "Sweepers"),
    "combi":                 ("Cleaning Equipment", "Floor Cleaning", "Combi Machines"),
}


@dataclass
class ParsedRow:
    row_number: int
    name_en: str
    model_number: str | None = None
    brand_name: str | None = None
    category_key: str = "accessories"
    # Set by the generic fallback parser for a free-text category read
    # straight from the sheet; when present it wins over category_key /
    # _CATEGORY_PATHS, since it isn't one of the fixed DK LAO taxonomy keys.
    category_free_text: str | None = None
    description_en: str | None = None
    specs: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0 and bool(self.name_en and self.name_en.strip())


# ── Helper: parse range strings like "2300 – 7000 mm" ────────────────────────

def _parse_range(raw: str) -> tuple[str, str, str | None]:
    """Return (min_val, max_val, unit) from a string like '2300 – 7000 mm'."""
    raw = raw.strip()
    unit: str | None = None
    for u in ["mm", " t", " V", "V", "t"]:
        if raw.endswith(u):
            raw = raw[: -len(u)].strip()
            unit = u.strip()
            break
    separators = [" – ", " - ", "–", "-", " to "]
    for sep in separators:
        if sep in raw:
            parts = raw.split(sep, 1)
            return parts[0].strip(), parts[1].strip(), unit
    return raw.strip(), raw.strip(), unit


def _cell(ws, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def _is_header_or_empty(text: str) -> bool:
    """Skip rows that appear to be headers or completely empty."""
    lower = text.lower()
    skip_keywords = {
        "model", "type", "ລຸ້ນ", "ປະເພດ", "ໝວດ", "ລາຍການ",
        "brand", "description", "ລາວ", "english", "ລາຍລະອຽດ",
        "ຄວາມສູງ", "ນ້ຳໜັກ", "ແຫຼ່ງ", "ໝາຍເຫດ",
    }
    return not text or any(kw in lower for kw in skip_keywords)


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def _parse_jungheinrich(ws) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    max_row = ws.max_row
    for r in range(3, max_row + 1):
        model = _cell(ws, r, 1)
        if not model or _is_header_or_empty(model):
            continue
        origin = _cell(ws, r, 2)
        lift_raw = _cell(ws, r, 3)
        cap_raw = _cell(ws, r, 4)
        volt_raw = _cell(ws, r, 5)

        parsed = ParsedRow(
            row_number=r,
            name_en=model,
            model_number=model,
            brand_name="Jungheinrich",
            category_key="jungheinrich_forklift",
        )

        if lift_raw:
            lo, hi, unit = _parse_range(lift_raw)
            parsed.specs.append({"group": "Performance", "key": "lift_height_min", "label": "Lift Height Min", "value": lo, "unit": unit or "mm"})
            parsed.specs.append({"group": "Performance", "key": "lift_height_max", "label": "Lift Height Max", "value": hi, "unit": unit or "mm"})
        if cap_raw:
            lo, hi, unit = _parse_range(cap_raw)
            parsed.specs.append({"group": "Performance", "key": "capacity_min", "label": "Capacity Min", "value": lo, "unit": unit or "t"})
            parsed.specs.append({"group": "Performance", "key": "capacity_max", "label": "Capacity Max", "value": hi, "unit": unit or "t"})
        if volt_raw:
            v = re.sub(r"[^\d.]", "", volt_raw)
            parsed.specs.append({"group": "Performance", "key": "voltage", "label": "Voltage", "value": v, "unit": "V"})
        if origin:
            parsed.specs.append({"group": "Origin", "key": "country_of_manufacture", "label": "Country of Manufacture", "value": origin, "unit": None})

        rows.append(parsed)
    return rows


def _parse_mitsubishi(ws) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    max_row = ws.max_row
    _TYPE_TO_CAT = {
        "diesel":    "mitsubishi_diesel",
        "lpg":       "mitsubishi_lpg",
        "3 wheel":   "mitsubishi_electric3",
        "4 wheel":   "mitsubishi_electric4",
        "electric":  "mitsubishi_electric3",
        "pallet":    "mitsubishi_other",
        "stacker":   "mitsubishi_other",
        "reach":     "mitsubishi_other",
        "order":     "mitsubishi_other",
    }
    for r in range(3, max_row + 1):
        type_name = _cell(ws, r, 1)
        if not type_name or _is_header_or_empty(type_name):
            continue
        lift_raw = _cell(ws, r, 2)
        cap_raw = _cell(ws, r, 3)
        volt_raw = _cell(ws, r, 4)

        cat_key = "mitsubishi_other"
        lower = type_name.lower()
        for keyword, key in _TYPE_TO_CAT.items():
            if keyword in lower:
                cat_key = key
                break

        parsed = ParsedRow(
            row_number=r,
            name_en=type_name,
            model_number=None,
            brand_name="Mitsubishi Forklift",
            category_key=cat_key,
        )

        if lift_raw:
            lo, hi, unit = _parse_range(lift_raw)
            parsed.specs.append({"group": "Performance", "key": "lift_height_min", "label": "Lift Height Min", "value": lo, "unit": unit or "mm"})
            parsed.specs.append({"group": "Performance", "key": "lift_height_max", "label": "Lift Height Max", "value": hi, "unit": unit or "mm"})
        if cap_raw:
            lo, hi, unit = _parse_range(cap_raw)
            parsed.specs.append({"group": "Performance", "key": "capacity_min", "label": "Capacity Min", "value": lo, "unit": unit or "t"})
            parsed.specs.append({"group": "Performance", "key": "capacity_max", "label": "Capacity Max", "value": hi, "unit": unit or "t"})
        if volt_raw and volt_raw.strip() != "-":
            v = re.sub(r"[^\d.]", "", volt_raw)
            if v:
                parsed.specs.append({"group": "Performance", "key": "voltage", "label": "Voltage", "value": v, "unit": "V"})

        rows.append(parsed)
    return rows


def _parse_accessories(ws) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    max_row = ws.max_row
    for r in range(3, max_row + 1):
        category_hint = _cell(ws, r, 1)
        item = _cell(ws, r, 2)
        notes = _cell(ws, r, 3)
        if not item or _is_header_or_empty(item):
            continue
        cat_key = "accessories"
        parsed = ParsedRow(
            row_number=r,
            name_en=item,
            model_number=None,
            brand_name=None,
            category_key=cat_key,
            description_en=notes or None,
        )
        if category_hint:
            parsed.specs.append({"group": "General", "key": "subcategory", "label": "Sub-category", "value": category_hint, "unit": None})
        rows.append(parsed)
    return rows


def _parse_warehouse(ws) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    max_row = ws.max_row
    _RACKING_KEYWORDS = {"racking", "shelving", "cantilever", "mezzanine", "storage", "shuttle", "high bay", "lift system"}
    for r in range(3, max_row + 1):
        cat_hint = _cell(ws, r, 1)
        name_lo = _cell(ws, r, 2)
        name_en = _cell(ws, r, 3)
        if not name_en or _is_header_or_empty(name_en):
            continue
        lower_hint = cat_hint.lower()
        cat_key = "racking" if any(k in lower_hint for k in _RACKING_KEYWORDS) else "warehouse"
        parsed = ParsedRow(
            row_number=r,
            name_en=name_en,
            model_number=None,
            brand_name=None,
            category_key=cat_key,
        )
        if name_lo:
            parsed.specs.append({"group": "General", "key": "name_lo", "label": "Lao Name", "value": name_lo, "unit": None})
        rows.append(parsed)
    return rows


def _parse_nilfisk_vacuum(ws) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    max_row = ws.max_row
    for r in range(3, max_row + 1):
        tier = _cell(ws, r, 1)
        model = _cell(ws, r, 2)
        desc = _cell(ws, r, 3)
        if not model or _is_header_or_empty(model):
            continue
        cat_key = "vacuum_industrial" if "industrial" in tier.lower() else "vacuum_commercial"
        parsed = ParsedRow(
            row_number=r,
            name_en=model,
            model_number=model,
            brand_name="Nilfisk",
            category_key=cat_key,
            description_en=desc or None,
        )
        if tier:
            parsed.specs.append({"group": "Technical", "key": "product_tier", "label": "Product Tier", "value": tier, "unit": None})
        rows.append(parsed)
    return rows


def _parse_nilfisk_floor(ws) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    max_row = ws.max_row
    _TYPE_TO_CAT = {
        "scrubber": "scrubber_dryer",
        "sweeper":  "sweeper",
        "combi":    "combi",
    }
    for r in range(3, max_row + 1):
        machine_type = _cell(ws, r, 1)
        model = _cell(ws, r, 2)
        desc = _cell(ws, r, 3)
        if not model or _is_header_or_empty(model):
            continue
        cat_key = "scrubber_dryer"
        lower = machine_type.lower()
        for keyword, key in _TYPE_TO_CAT.items():
            if keyword in lower:
                cat_key = key
                break
        parsed = ParsedRow(
            row_number=r,
            name_en=model,
            model_number=model,
            brand_name="Nilfisk",
            category_key=cat_key,
            description_en=desc or None,
        )
        if machine_type:
            parsed.specs.append({"group": "Technical", "key": "machine_type", "label": "Machine Type", "value": machine_type, "unit": None})
        rows.append(parsed)
    return rows


def _parse_brands_sheet(ws) -> list[ParsedRow]:
    """Brand sheet: extract brand names only — no products."""
    return []


# ── Generic fallback parser ───────────────────────────────────────────────────
# Used only when NO sheet in the workbook matches a known DK LAO template
# keyword (see _detect_handler). Rather than silently importing nothing, this
# reads the first sheet with a flexible, header-driven column mapping —
# case-insensitive and whitespace-tolerant, like the Inventory importer.

_GENERIC_HEADER_ALIASES: dict[str, str] = {
    "name": "name_en", "product": "name_en", "product_name": "name_en",
    "item": "name_en", "item_name": "name_en", "name_en": "name_en",
    "model": "model_number", "model_number": "model_number", "model_no": "model_number", "sku": "model_number",
    "brand": "brand_name", "brand_name": "brand_name", "manufacturer": "brand_name",
    "category": "category_l1", "category_l1": "category_l1", "type": "category_l1",
    "description": "description_en", "desc": "description_en",
    "notes": "description_en", "remark": "description_en", "remarks": "description_en",
}


def _normalize_generic_header(raw: str) -> str | None:
    key = raw.strip().lower().replace(" ", "_").replace("-", "_")
    return _GENERIC_HEADER_ALIASES.get(key)


def _find_header_row(ws, max_scan: int = 5) -> tuple[int, dict[int, str]] | None:
    """Scan the first few rows for whichever one looks most like a header —
    the row with the most cells matching a known alias. Returns None if no
    row within the scan window contains a recognizable 'name' column."""
    best_row: int | None = None
    best_map: dict[int, str] = {}
    best_score = 0
    max_col = ws.max_column or 1
    for r in range(1, max_scan + 1):
        col_map: dict[int, str] = {}
        for c in range(1, max_col + 1):
            raw = _cell(ws, r, c)
            if not raw:
                continue
            canon = _normalize_generic_header(raw)
            if canon and canon not in col_map.values():
                col_map[c] = canon
        if len(col_map) > best_score:
            best_score = len(col_map)
            best_row = r
            best_map = col_map
    if best_row is None or "name_en" not in best_map.values():
        return None
    return best_row, best_map


def _parse_generic(ws) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    header = _find_header_row(ws)
    if header is None:
        logger.warning(
            "Generic fallback parser: no recognizable header row (with a Name/Product column) "
            "found in the first %d rows of sheet %r — nothing to import.",
            5, ws.title,
        )
        return rows

    header_row, col_map = header
    logger.info(
        "Generic fallback parser: sheet %r — detected header row %d with columns %s",
        ws.title, header_row, {c: canon for c, canon in col_map.items()},
    )

    max_row = ws.max_row
    for r in range(header_row + 1, max_row + 1):
        values: dict[str, str] = {}
        for c, canon in col_map.items():
            text = _cell(ws, r, c)
            if text:
                values[canon] = text

        name_en = values.get("name_en", "")
        if not name_en:
            continue  # blank row — not an error, just nothing to import

        parsed = ParsedRow(
            row_number=r,
            name_en=name_en,
            model_number=values.get("model_number") or None,
            brand_name=values.get("brand_name") or None,
            category_free_text=values.get("category_l1") or None,
            description_en=values.get("description_en") or None,
        )
        rows.append(parsed)

    logger.info(
        "Generic fallback parser: sheet %r — parsed %d data row(s) out of %d total row(s) scanned.",
        ws.title, len(rows), max(max_row - header_row, 0),
    )
    return rows


_HANDLERS: dict[str, Any] = {
    "jungheinrich_forklift": _parse_jungheinrich,
    "mitsubishi_forklift":   _parse_mitsubishi,
    "accessories":           _parse_accessories,
    "warehouse":             _parse_warehouse,
    "vacuum":                _parse_nilfisk_vacuum,
    "floor_cleaner":         _parse_nilfisk_floor,
    "brands":                _parse_brands_sheet,
}


def _detect_handler(sheet_name: str) -> str | None:
    lower = sheet_name.lower()
    if sheet_name in _IGNORED_SHEETS:
        return None
    for keyword, handler in _SHEET_MAP:
        if keyword.lower() in lower:
            return handler
    return None


# ── Main service ──────────────────────────────────────────────────────────────

class ImportService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db
        self._import_repo = ImportRepository(db)
        self._product_repo = ProductRepository(db)
        self._brand_svc = BrandService(db)
        self._category_svc = CategoryService(db)

    async def _rows_to_preview(
        self, parsed_rows: list[ParsedRow], sheet_name: str,
    ) -> tuple[list[PreviewProduct], list[PreviewError]]:
        """Shared by both the keyword-matched handlers and the generic
        fallback parser — converts ParsedRow objects into preview rows,
        resolving the create/update action and category path for each."""
        valid_rows: list[PreviewProduct] = []
        error_rows: list[PreviewError] = []

        for row in parsed_rows:
            if row.is_valid:
                if row.category_free_text:
                    # Free-text category from the generic parser — not one
                    # of the fixed DK LAO taxonomy keys, so it bypasses
                    # _CATEGORY_PATHS and is used (and auto-created) as-is.
                    cat = (row.category_free_text, None, None)
                else:
                    cat = _CATEGORY_PATHS.get(row.category_key, (None, None, None))
                action = "create"
                if row.model_number:
                    existing = await self._product_repo.get_by_model_and_brand(
                        row.model_number, None
                    )
                    if existing:
                        action = "update"
                valid_rows.append(PreviewProduct(
                    row_number=row.row_number,
                    action=action,
                    name_en=row.name_en,
                    model_number=row.model_number,
                    brand_name=row.brand_name,
                    category_l1=cat[0],
                    category_l2=cat[1],
                    category_l3=cat[2],
                    description_en=row.description_en,
                    specs=[
                        PreviewSpec(
                            spec_group=s["group"],
                            spec_key=s["key"],
                            spec_label=s["label"],
                            spec_value=s["value"],
                            spec_unit=s.get("unit"),
                        )
                        for s in row.specs
                    ],
                ))
            else:
                for err in row.errors:
                    error_rows.append(PreviewError(
                        row_number=row.row_number,
                        sheet_name=sheet_name,
                        error_message=err,
                        row_data={"name_en": row.name_en, "model": row.model_number},
                    ))

        return valid_rows, error_rows

    async def preview(
        self,
        file_content: bytes,
        filename: str,
        created_by: int,
        max_file_size: int = 10 * 1024 * 1024,
    ) -> ImportPreviewResponse:
        if len(file_content) > max_file_size:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="File exceeds the 10 MB limit.",
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                io.BytesIO(file_content), read_only=True, data_only=True
            )
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=f"Cannot open workbook: {exc}",
            )

        logger.info(
            "Import preview: workbook %r has %d sheet(s): %s",
            filename, len(wb.sheetnames), wb.sheetnames,
        )

        sheets_detected: list[str] = []
        sheet_previews: list[SheetPreview] = []
        total_valid = 0
        total_errors = 0

        for sheet_name in wb.sheetnames:
            handler_key = _detect_handler(sheet_name)
            if handler_key is None or handler_key == "brands":
                continue

            sheets_detected.append(sheet_name)
            ws = wb[sheet_name]
            parser = _HANDLERS.get(handler_key)
            if parser is None:
                continue

            try:
                parsed_rows = parser(ws)
            except Exception as exc:
                logger.error("Error parsing sheet %s: %s", sheet_name, exc)
                continue

            logger.info(
                "Sheet %r (handler=%s): parsed %d row(s) out of %d row(s) in the sheet.",
                sheet_name, handler_key, len(parsed_rows), ws.max_row,
            )

            valid_rows, error_rows = await self._rows_to_preview(parsed_rows, sheet_name)
            total_valid += len(valid_rows)
            total_errors += len(error_rows)

            sheet_previews.append(SheetPreview(
                sheet_name=sheet_name,
                handler=handler_key,
                valid_rows=valid_rows,
                error_rows=error_rows,
                total_valid=len(valid_rows),
                total_errors=len(error_rows),
            ))

        if not sheets_detected and wb.sheetnames:
            # No sheet name matched a known DK LAO template keyword — rather
            # than silently importing nothing, fall back to a flexible,
            # header-driven read of the first sheet.
            first_sheet_name = wb.sheetnames[0]
            ws = wb[first_sheet_name]
            logger.warning(
                "No sheet in %r matched a known template keyword (sheets: %s) — "
                "falling back to the generic parser on the first sheet %r (%d rows x %d cols).",
                filename, wb.sheetnames, first_sheet_name, ws.max_row, ws.max_column,
            )
            try:
                parsed_rows = _parse_generic(ws)
            except Exception as exc:
                logger.error("Generic fallback parser failed on sheet %r: %s", first_sheet_name, exc)
                parsed_rows = []

            sheets_detected.append(first_sheet_name)
            valid_rows, error_rows = await self._rows_to_preview(parsed_rows, first_sheet_name)
            total_valid += len(valid_rows)
            total_errors += len(error_rows)

            sheet_previews.append(SheetPreview(
                sheet_name=first_sheet_name,
                handler="generic",
                valid_rows=valid_rows,
                error_rows=error_rows,
                total_valid=len(valid_rows),
                total_errors=len(error_rows),
            ))

        logger.info(
            "Import preview: %r — total_valid=%d, total_errors=%d across %d sheet(s).",
            filename, total_valid, total_errors, len(sheets_detected),
        )

        wb.close()

        # Persist the job so the user can execute later
        preview_payload = {
            "sheets": [sp.model_dump() for sp in sheet_previews],
            "total_valid": total_valid,
            "total_errors": total_errors,
        }
        job = ImportJob(
            original_filename=filename,
            status=ImportStatus.PREVIEW.value,
            total_rows=total_valid + total_errors,
            processed_rows=0,
            success_rows=0,
            error_rows=total_errors,
            preview_data=preview_payload,
            created_by=created_by,
        )
        job = await self._import_repo.create(job)
        await self.db.commit()

        return ImportPreviewResponse(
            job_id=job.id,
            filename=filename,
            sheets_detected=sheets_detected,
            total_valid=total_valid,
            total_errors=total_errors,
            sheets=sheet_previews,
        )

    async def execute(self, job_id: int, executed_by: int) -> ImportExecuteResponse:
        job = await self._import_repo.get_by_id(job_id)
        if job is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Import job {job_id} not found.",
            )
        if job.status not in (ImportStatus.PREVIEW.value, ImportStatus.FAILED.value):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Job is in '{job.status}' state; only PREVIEW jobs can be executed.",
            )

        await self._import_repo.update(
            job, {"status": ImportStatus.PROCESSING.value}
        )
        await self.db.commit()

        preview_data = job.preview_data or {}
        sheets: list[dict] = preview_data.get("sheets", [])

        success = 0
        errors: list[PreviewError] = []

        for sheet in sheets:
            sheet_name = sheet.get("sheet_name", "")
            for row_data in sheet.get("valid_rows", []):
                row_num = row_data.get("row_number", 0)
                try:
                    await self._upsert_product(row_data, executed_by)
                    success += 1
                except Exception as exc:
                    logger.error("Import error row %s sheet %s: %s", row_num, sheet_name, exc)
                    err_msg = str(exc)
                    errors.append(PreviewError(
                        row_number=row_num,
                        sheet_name=sheet_name,
                        error_message=err_msg,
                    ))
                    err = ImportError(
                        job_id=job.id,
                        sheet_name=sheet_name,
                        row_number=row_num,
                        error_message=err_msg,
                    )
                    await self._import_repo.add_error(err)

        final_status = ImportStatus.COMPLETED.value
        if errors and success == 0:
            final_status = ImportStatus.FAILED.value
        elif errors:
            final_status = ImportStatus.PARTIAL.value

        await self._import_repo.update(job, {
            "status": final_status,
            "processed_rows": success + len(errors),
            "success_rows": success,
            "error_rows": len(errors),
            "completed_at": datetime.now(UTC),
            "preview_data": None,  # free memory after execution
        })
        await self.db.commit()

        if success + len(errors) > 0:
            # Enterprise Audit & Alert: one summary notification for the whole
            # catalog import batch, not per-row.
            try:
                actor_name = await resolve_actor_name(self.db, executed_by)
                await NotificationService(self.db).notify_role(
                    role=ADMIN_ROLE,
                    subject="Catalog Import Completed",
                    message=(
                        f"{actor_name} imported the product catalog: "
                        f"{success} product(s) created/updated, {len(errors)} row error(s)."
                    ),
                    event_type="catalog.import_completed",
                    entity_type="catalog_import",
                    entity_id=job.id,
                )
            except Exception:
                logger.exception("Failed to notify admins of catalog import completion")

        return ImportExecuteResponse(
            job_id=job.id,
            status=final_status,
            success_rows=success,
            error_rows=len(errors),
            errors=errors,
        )

    async def _upsert_product(self, row_data: dict, created_by: int) -> None:
        name_en: str = row_data.get("name_en", "").strip()
        model_number: str | None = row_data.get("model_number")
        brand_name: str | None = row_data.get("brand_name")
        description_en: str | None = row_data.get("description_en")
        category_l1: str | None = row_data.get("category_l1")
        category_l2: str | None = row_data.get("category_l2")
        category_l3: str | None = row_data.get("category_l3")
        raw_specs: list[dict] = row_data.get("specs", [])

        if not name_en:
            raise ValueError("Product name_en is empty")

        # Resolve brand
        brand_id: int | None = None
        if brand_name:
            brand = await self._brand_svc.get_or_create_by_name(brand_name)
            brand_id = brand.id

        # Resolve category
        category_id: int | None = None
        cat = await self._category_svc.get_or_create_by_path(
            category_l1, category_l2, category_l3
        )
        if cat:
            category_id = cat.id

        # Upsert by (model_number, brand_id)
        existing = None
        if model_number and brand_id:
            existing = await self._product_repo.get_by_model_and_brand(model_number, brand_id)

        if existing:
            # Update: refresh name, description, category only; do not overwrite flags
            changes: dict = {
                "name_en": name_en,
                "category_id": category_id,
                "updated_by": created_by,
            }
            if description_en:
                changes["description_en"] = description_en
            await self._product_repo.update(existing, changes)
            product_id = existing.id
            # Re-seed specs (wipe per-group, then re-insert)
            for spec in raw_specs:
                await self._product_repo.clear_specs_by_group(product_id, spec["group"])
        else:
            # Create
            product_svc = ProductService(self.db)
            new_product = await product_svc.create_product(
                ProductCreate(
                    name_en=name_en,
                    model_number=model_number,
                    brand_id=brand_id,
                    category_id=category_id,
                    description_en=description_en,
                ),
                created_by=created_by,
            )
            product_id = new_product.id

        # Upsert specs
        from app.models.product import ProductSpec
        for s in raw_specs:
            try:
                spec_key = re.sub(r"[^a-z0-9_]", "_", s["key"].lower())
                if not re.match(r"^[a-z]", spec_key):
                    spec_key = "s_" + spec_key
                spec = ProductSpec(
                    product_id=product_id,
                    spec_group=s["group"],
                    spec_key=spec_key,
                    spec_label=s["label"],
                    spec_value=str(s["value"]),
                    spec_unit=s.get("unit"),
                    sort_order=0,
                )
                self.db.add(spec)
            except Exception as exc:
                logger.warning("Skipping invalid spec row: %s", exc)

        await self.db.flush()
